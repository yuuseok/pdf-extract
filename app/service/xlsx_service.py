import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class XlsxService:
    """Excel(.xlsx) 파일에서 text, markdown, json을 추출."""

    def extract(self, file_path: str) -> dict:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        elements = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                str_row = [str(cell) if cell is not None else "" for cell in row]
                if any(c for c in str_row):  # 빈 행 제외
                    rows.append(str_row)

            if rows:
                elements.append({
                    "type": "sheet",
                    "sheet_name": sheet_name,
                    "rows": rows,
                    "row_count": len(rows),
                    "col_count": len(rows[0]) if rows else 0,
                })

        wb.close()

        text = self._to_text(elements)
        markdown = self._to_markdown(elements)

        return {
            "text": text,
            "markdown": markdown,
            "json": elements,
            "json_raw": elements,
            "page_count": len(elements),  # 시트 수
        }

    def _to_text(self, elements: list) -> str:
        lines = []
        for el in elements:
            lines.append(f"[{el['sheet_name']}]")
            for row in el["rows"]:
                lines.append("\t".join(row))
            lines.append("")
        return "\n".join(lines)

    def _to_markdown(self, elements: list) -> str:
        lines = []
        for el in elements:
            lines.append(f"## {el['sheet_name']}")
            lines.append("")
            rows = el["rows"]
            if not rows:
                continue

            # 첫 행을 헤더로
            col_count = max(len(r) for r in rows)
            header = rows[0] + [""] * (col_count - len(rows[0]))
            lines.append("| " + " | ".join(header) + " |")
            lines.append("| " + " | ".join(["---"] * col_count) + " |")

            for row in rows[1:]:
                padded = row + [""] * (col_count - len(row))
                lines.append("| " + " | ".join(padded) + " |")
            lines.append("")
        return "\n".join(lines)
